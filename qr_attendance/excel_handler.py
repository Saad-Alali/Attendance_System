import pandas as pd
import numpy as np
from datetime import datetime
import os
import difflib
import openpyxl
from openpyxl.styles import numbers
from qr_attendance.config import COLUMN_NAMES, ATTENDANCE_STATUS, AUTHORIZED_ABSENCE

class ExcelHandler:
    def __init__(self, file_path=None):
        self.file_path = file_path
        self.df = None
        self.present_students = set()
        self.original_formats = {}
        self.attendance_col_idx = None
        self.original_workbook = None
        self.date_col_idx = None
        self.text_format_columns = {
            "attendance": None,    # مؤشر الحضور
            "expected_hours": None,  # الساعات المتوقعة
            "actual_hours": None,    # الساعات الفعلية
            "absence_hours": None,   # ساعات الغياب
        }
        
    def load_file(self, file_path=None):
        if file_path:
            self.file_path = file_path
        
        if not self.file_path or not os.path.exists(self.file_path):
            raise FileNotFoundError("Excel file not found")
        
        # Load with openpyxl first to store the original format and file structure
        try:
            self.original_workbook = openpyxl.load_workbook(self.file_path)
            self._store_column_indices()
        except Exception as e:
            print(f"Warning: Could not load original workbook with openpyxl: {e}")
        
        # Now load with pandas for data manipulation - keep all text columns as string
        dtype_dict = {}
        for key in self.text_format_columns.keys():
            col_name = COLUMN_NAMES.get(key)
            if col_name:
                dtype_dict[col_name] = str
        
        # Load with specific dtypes to preserve text format
        self.df = pd.read_excel(self.file_path, dtype=dtype_dict)
        
        # Convert necessary columns to string
        self._convert_columns_to_string()
        
        return self.df
    
    def _store_column_indices(self):
        """Store indices of important columns"""
        if not self.original_workbook:
            return
            
        ws = self.original_workbook.active
        header_row = 1
        
        # Find important column indices
        for idx, cell in enumerate(ws[header_row], 1):
            cell_value = cell.value
            
            # Store text format column indices
            for key, col_name in COLUMN_NAMES.items():
                if cell_value == col_name and key in self.text_format_columns:
                    self.text_format_columns[key] = idx
                    print(f"Found {key} column '{col_name}' at index {idx}")
            
            # Store date column index
            if cell_value == COLUMN_NAMES["date"]:
                self.date_col_idx = idx
                print(f"Found date column '{cell_value}' at index {idx}")
    
    def _convert_columns_to_string(self):
        """Convert specified columns to string format and replace NaN with empty string"""
        # Get all text format column names
        columns_to_convert = [COLUMN_NAMES[key] for key in self.text_format_columns.keys() 
                              if COLUMN_NAMES.get(key) in self.df.columns]
        
        for col in columns_to_convert:
            if col in self.df.columns:
                try:
                    # Convert to string
                    temp_col = self.df[col].astype(str)
                    # Replace NaN values with empty string
                    temp_col = temp_col.replace('nan', '')
                    self.df[col] = temp_col
                    print(f"Converted column '{col}' to string type")
                except Exception as e:
                    print(f"Warning: Could not convert column '{col}': {e}")
    
    def convert_date_column(self, date_column=None):
        """Convert date column to datetime format"""
        if self.df is None:
            raise ValueError("Excel file must be loaded first")
        
        if date_column is None:
            date_column = COLUMN_NAMES["date"]
        
        if date_column in self.df.columns:
            try:
                # Check if we're dealing with Excel numeric dates
                if pd.api.types.is_numeric_dtype(self.df[date_column]):
                    # Convert Excel numeric dates (days since 1899-12-30) to datetime
                    base_date = pd.Timestamp('1899-12-30')
                    days = pd.to_timedelta(self.df[date_column], unit='D')
                    self.df[date_column] = base_date + days
                    print(f"Converted Excel numeric dates in '{date_column}' to datetime")
                else:
                    # Regular date conversion for string dates
                    self.df[date_column] = pd.to_datetime(self.df[date_column], errors='coerce')
                    print(f"Converted string dates in '{date_column}' to datetime")
            except Exception as e:
                print(f"Error during date conversion: {e}")
                raise ValueError(f"Error converting date column: {e}")
        else:
            raise ValueError(f"Column {date_column} not found in file")
        
        return self.df
    
    def check_lecture_today(self, date_column=None):
        """Check if there's a lecture scheduled for today"""
        if self.df is None:
            raise ValueError("Excel file must be loaded first")
        
        if date_column is None:
            date_column = COLUMN_NAMES["date"]
        
        # Make sure date column exists
        if date_column not in self.df.columns:
            raise ValueError(f"Column {date_column} not found in file")
        
        # Make sure the date column is in datetime format
        if not pd.api.types.is_datetime64_any_dtype(self.df[date_column]):
            self.convert_date_column(date_column)
        
        today = datetime.now().date()
        
        # Find lectures scheduled for today
        today_lectures = self.df[self.df[date_column].dt.date == today]
        
        if today_lectures.empty:
            return None
        
        return today_lectures
    
    def _find_best_name_match(self, input_name, date_filter):
        """Find the best matching student name for a given date"""
        full_name_col = COLUMN_NAMES["full_name"]
        date_col = COLUMN_NAMES["date"]
        
        input_name = ' '.join(input_name.split())
        
        today_students = self.df[self.df[date_col].dt.date == date_filter.date()]
        if today_students.empty:
            return None, 0
        
        best_score = 0
        best_match = None
        
        for idx, row in today_students.iterrows():
            db_name = row[full_name_col]
            if not isinstance(db_name, str):
                db_name = str(db_name)
                
            db_name = ' '.join(db_name.split())
                
            similarity = difflib.SequenceMatcher(None, input_name.lower(), db_name.lower()).ratio()
            
            if similarity > best_score and similarity >= 0.8:
                best_score = similarity
                best_match = db_name
                
        return best_match, best_score
    
    def mark_attendance(self, student_name, student_id, lecture_date):
        """Mark a student as present"""
        if self.df is None:
            raise ValueError("Excel file must be loaded first")
        
        full_name_col = COLUMN_NAMES["full_name"]
        student_id_col = COLUMN_NAMES["student_id"]
        date_col = COLUMN_NAMES["date"]
        attendance_col = COLUMN_NAMES["attendance"]
        expected_hours_col = COLUMN_NAMES["expected_hours"]
        actual_hours_col = COLUMN_NAMES["actual_hours"]
        absence_hours_col = COLUMN_NAMES["absence_hours"]
        
        student_name = ' '.join(student_name.split())
        student_id = str(student_id).strip()
        
        best_name_match, match_score = self._find_best_name_match(student_name, lecture_date)
        
        student_mask = (
            (self.df[student_id_col].astype(str).str.strip() == student_id) &
            (self.df[date_col].dt.date == lecture_date.date())
        )
        
        if any(student_mask):
            student_idx = student_mask.idxmax()
            if isinstance(student_idx, pd.Series):
                student_idx = student_idx.iloc[0]
                
            found_name = self.df.loc[student_idx, full_name_col]
            found_name = ' '.join(str(found_name).split())
            
            print(f"Found match by ID. Database name: {found_name}, Entered name: {student_name}")
            similarity = difflib.SequenceMatcher(None, student_name.lower(), found_name.lower()).ratio()
            print(f"Name similarity: {similarity:.2f}")
            
            if similarity >= 0.8:
                # Convert all required columns to string (if needed)
                self._convert_columns_to_string()
                
                # Mark as present
                attendance_value = str(ATTENDANCE_STATUS["present"])
                
                for idx in self.df.index[student_mask]:
                    # Set attendance column to "حاضر"
                    self.df.at[idx, attendance_col] = attendance_value
                
                # Copy expected hours to actual hours, and clear absence hours
                if expected_hours_col in self.df.columns and actual_hours_col in self.df.columns:
                    for idx in self.df.index[student_mask]:
                        expected_value = self.df.at[idx, expected_hours_col]
                        
                        # Set actual hours - keep as string
                        self.df.at[idx, actual_hours_col] = expected_value
                        
                        # Clear absence hours
                        if absence_hours_col in self.df.columns:
                            self.df.at[idx, absence_hours_col] = ""
                
                for idx in self.df.index[student_mask]:
                    db_name = self.df.loc[idx, full_name_col]
                    student_key = (db_name, student_id, lecture_date.date())
                    self.present_students.add(student_key)
                
                return True, "Attendance recorded successfully"
            else:
                return False, f"Student name doesn't match. Did you mean {found_name}? Please enter the correct name."
        elif best_name_match and match_score >= 0.8:
            return False, f"Student ID incorrect. Did you mean {best_name_match}? Please enter the correct ID."
        else:
            return False, "Student not registered for this lecture"
    
    def mark_all_absent(self, lecture_date):
        """Mark students as absent if not present"""
        if self.df is None:
            raise ValueError("Excel file must be loaded first")
        
        full_name_col = COLUMN_NAMES["full_name"]
        student_id_col = COLUMN_NAMES["student_id"]
        date_col = COLUMN_NAMES["date"]
        attendance_col = COLUMN_NAMES["attendance"]
        expected_hours_col = COLUMN_NAMES["expected_hours"]
        actual_hours_col = COLUMN_NAMES["actual_hours"]
        absence_hours_col = COLUMN_NAMES["absence_hours"]
        authorized_absence_col = COLUMN_NAMES["authorized_absence"]
        
        date_mask = self.df[date_col].dt.date == lecture_date.date()
        today_indices = self.df.index[date_mask].tolist()
        
        absent_indices = []
        
        for idx in today_indices:
            student_name = self.df.at[idx, full_name_col]
            student_id = self.df.at[idx, student_id_col]
            student_key = (student_name, str(student_id), lecture_date.date())
            
            if student_key not in self.present_students:
                absent_indices.append(idx)
        
        if not absent_indices:
            return 0
        
        # Convert all required columns to string (if needed)
        self._convert_columns_to_string()
        
        # Mark absent students
        absent_value = str(ATTENDANCE_STATUS["absent"])
        
        for idx in absent_indices:
            # Set attendance column to "غائب"
            self.df.at[idx, attendance_col] = absent_value
            
            # Handle hours - copy expected to absence, clear actual
            if expected_hours_col in self.df.columns:
                expected_value = self.df.at[idx, expected_hours_col]
                
                # Set absence hours - keep as string
                if absence_hours_col in self.df.columns:
                    self.df.at[idx, absence_hours_col] = expected_value
                
                # Clear actual hours
                if actual_hours_col in self.df.columns:
                    self.df.at[idx, actual_hours_col] = ""
            
            # Set authorized absence to "لا"
            if authorized_absence_col in self.df.columns:
                self.df.at[idx, authorized_absence_col] = AUTHORIZED_ABSENCE["no"]
        
        return len(absent_indices)
    
    def save_file(self, output_path=None):
        """
        Save Excel file while ensuring proper formatting for text columns
        """
        if self.df is None:
            raise ValueError("Excel file must be loaded first")
        
        if output_path is None:
            output_path = self.file_path
        
        try:
            # Create backup of original file
            if os.path.exists(self.file_path):
                backup_path = f"{self.file_path}.bak"
                try:
                    import shutil
                    shutil.copy2(self.file_path, backup_path)
                    print(f"Backup created at: {backup_path}")
                except Exception as e:
                    print(f"Warning: Could not create backup: {e}")
            
            # Save the DataFrame to a temporary CSV file
            temp_csv = f"{output_path}.temp.csv"
            self.df.to_csv(temp_csv, index=False, encoding='utf-8-sig')
            print(f"DataFrame saved to temporary CSV: {temp_csv}")
            
            # Load the original Excel file with openpyxl
            if self.original_workbook is None:
                # If we don't have the original workbook, try to load it
                try:
                    self.original_workbook = openpyxl.load_workbook(self.file_path)
                except Exception:
                    # If that fails, create a new workbook
                    self.original_workbook = openpyxl.Workbook()
            
            # Get the active worksheet
            ws = self.original_workbook.active
            
            # Load the temporary CSV into a new DataFrame
            updated_df = pd.read_csv(temp_csv, encoding='utf-8-sig')
            
            # Find headers
            header_row = 1
            column_indices = {}
            
            for idx, cell in enumerate(ws[header_row], 1):
                if cell.value in COLUMN_NAMES.values():
                    column_indices[cell.value] = idx
            
            # Get indices for special columns
            date_idx = column_indices.get(COLUMN_NAMES['date'])
            
            # Map column names to their indices
            text_column_indices = {}
            for key, col_name in COLUMN_NAMES.items():
                if key in self.text_format_columns and col_name in column_indices:
                    text_column_indices[col_name] = column_indices[col_name]
            
            # Create mapping of all columns to their indices
            column_mapping = {}
            for i, col_name in enumerate(updated_df.columns):
                if col_name in column_indices:
                    column_mapping[i] = column_indices[col_name]
            
            # Update values in the worksheet
            for row_idx, row in enumerate(updated_df.values, 2):  # Start from row 2 (skip header)
                for col_idx, value in enumerate(row):
                    if col_idx in column_mapping:
                        excel_col_idx = column_mapping[col_idx]
                        cell = ws.cell(row=row_idx, column=excel_col_idx)
                        
                        # Get column name
                        col_name = updated_df.columns[col_idx]
                        
                        # Special handling for text format columns (attendance, hours, etc.)
                        if col_name in text_column_indices:
                            # Handle null/NaN values
                            if pd.isna(value) or value == 'nan' or value is None or value == '':
                                cell.value = ''
                            else:
                                # Use the value directly without any conversion
                                cell.value = str(value)
                            
                            # Set number format to TEXT
                            cell.number_format = '@'
                        
                        # Special handling for date column
                        elif date_idx and excel_col_idx == date_idx:
                            # If it's a date, format it as MM/DD/YYYY
                            if pd.notna(value) and value != '':
                                try:
                                    date_value = pd.to_datetime(value)
                                    # Store as date in Excel with custom format
                                    cell.value = date_value
                                    # Apply custom date format MM/DD/YYYY
                                    cell.number_format = 'M/D/YYYY'
                                except:
                                    # If conversion fails, just use the value as is
                                    cell.value = value
                            else:
                                cell.value = value
                        
                        else:
                            # For other columns, just update the value
                            cell.value = value
            
            # Ensure TEXT format for all text columns (loop through all text format columns)
            for col_name, col_idx in text_column_indices.items():
                if col_idx:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    for cell in ws[col_letter]:
                        cell.number_format = '@'
            
            # Ensure date format for entire date column
            if date_idx:
                col_letter = openpyxl.utils.get_column_letter(date_idx)
                for row_idx in range(2, len(self.df) + 2):  # Skip header
                    cell = ws.cell(row=row_idx, column=date_idx)
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'M/D/YYYY'
            
            # Save the updated workbook
            self.original_workbook.save(output_path)
            print(f"Excel file saved with formatted columns: {output_path}")
            
            # Clean up temporary file
            try:
                os.remove(temp_csv)
            except:
                pass
            
            return True, f"File saved successfully at {output_path}"
            
        except Exception as e:
            print(f"Error details: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Try fallback save method
            try:
                print("Attempting fallback save method...")
                self.df.to_excel(output_path, index=False)
                print(f"File saved using fallback method (without formatting) at {output_path}")
                return True, f"File saved using fallback method at {output_path}"
            except Exception as fallback_error:
                print(f"Fallback save failed: {fallback_error}")
                return False, f"Error saving file: {e}"
    
    def reset_attendance_for_date(self, lecture_date):
        """Reset attendance data for a specific date"""
        if self.df is None:
            raise ValueError("Excel file must be loaded first")
        
        date_col = COLUMN_NAMES["date"]
        attendance_col = COLUMN_NAMES["attendance"]
        absence_hours_col = COLUMN_NAMES["absence_hours"]
        authorized_absence_col = COLUMN_NAMES["authorized_absence"]
        actual_hours_col = COLUMN_NAMES["actual_hours"]
        
        date_mask = self.df[date_col].dt.date == lecture_date.date()
        date_indices = self.df.index[date_mask].tolist()
        
        if not date_indices:
            return 0
        
        for idx in date_indices:
            # Clear all attendance-related columns
            if attendance_col in self.df.columns:
                self.df.at[idx, attendance_col] = ""
            
            if absence_hours_col in self.df.columns:
                self.df.at[idx, absence_hours_col] = ""
            
            if authorized_absence_col in self.df.columns:
                self.df.at[idx, authorized_absence_col] = ""

            if actual_hours_col in self.df.columns:
                self.df.at[idx, actual_hours_col] = ""
        
        # Clear present students set for this date
        self.present_students = {key for key in self.present_students 
                               if key[2] != lecture_date.date()}
        
        return len(date_indices)